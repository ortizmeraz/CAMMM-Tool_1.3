import openpyxl
from openpyxl import load_workbook

# MapName='Closnes-Quebec-Direct'
# MapCode='cki95ew494au719ufyzwo9v7j'
# TileSet='quebec-busnetwork'
# CoordMapX='-71.29481330370135'
# CoordMapY='46.81042223815351'
# URLtileset='carmela-cucuzzella.ckfr7loqt01ml23p7okf6ir0n-13m3b'
# SourceLayerId='quebec-busnetwork' 
# SourceLayerName='Quebec_BusNetwork'
# Variable='CatClossnes'
# OneColorH='144'
# OneColorS='96'
# OneColorL='62'
# TwoColorH='144'
# TwoColorS='92'
# TwoColorL='25'


# 3 name
# 86 id-map (it is self generated once uploaded)
# 120 tile name
# 203 point coordinates
# 209 Tileset ID
 
# 2908 id name
# 2911 Source layer name

# 2926 mapping category 
# 2928 color start

# 2930 color end

# 2935 mapping category

# 2946 id-map (it is self generated once uploaded)



# 64-10999
# 128-11999



# MapName='Test for now'
# MapCode='cki95ew494ajijijijijijijiji'
# TileSet='Torino_BusMetwork'
# CoordMapX='7.6945'
# CoordMapY='45.0617'
# URLtileset='carmela-cucuzzella.ckfr7o7ko01do23mpzo5wf3iz-63aps'
# SourceLayerId='Torino_BusMetwork' 
# SourceLayerName='Torino_BusMetwork'
# Variable='CatClossnes_11111'
# OneColorH='144'
# OneColorS='96'
# OneColorL='62'
# TwoColorH='144'
# TwoColorS='92'
# TwoColorL='25'



def TextSource():
    Text='''{
        "version": 8,
        "name": "'''+MapName+'''",
        "metadata": {
            "mapbox:type": "default",
            "mapbox:origin": "basic-v1",
            "mapbox:sdk-support": {
                "android": "9.3.0",
                "ios": "5.10.0",
                "js": "1.10.0"
            },
            "mapbox:autocomposite": true,
            "mapbox:groups": {
                "Transit, transit-labels": {
                    "name": "Transit, transit-labels",
                    "collapsed": false
                },
                "Administrative boundaries, admin": {
                    "name": "Administrative boundaries, admin",
                    "collapsed": false
                },
                "Land & water, built": {
                    "name": "Land & water, built",
                    "collapsed": false
                },
                "Land & water, land": {
                    "name": "Land & water, land",
                    "collapsed": false
                },
                "Road network, bridges": {
                    "name": "Road network, bridges",
                    "collapsed": false
                },
                "Road network, tunnels": {
                    "name": "Road network, tunnels",
                    "collapsed": false
                },
                "Road network, road-labels": {
                    "name": "Road network, road-labels",
                    "collapsed": false
                },
                "Buildings, built": {
                    "name": "Buildings, built",
                    "collapsed": false
                },
                "Natural features, natural-labels": {
                    "name": "Natural features, natural-labels",
                    "collapsed": false
                },
                "Road network, surface": {
                    "name": "Road network, surface",
                    "collapsed": false
                },
                "Walking, cycling, etc., barriers-bridges": {
                    "name": "Walking, cycling, etc., barriers-bridges",
                    "collapsed": false
                },
                "Place labels, place-labels": {
                    "name": "Place labels, place-labels",
                    "collapsed": false
                },
                "Point of interest labels, poi-labels": {
                    "name": "Point of interest labels, poi-labels",
                    "collapsed": false
                },
                "Walking, cycling, etc., tunnels": {
                    "name": "Walking, cycling, etc., tunnels",
                    "collapsed": false
                },
                "Walking, cycling, etc., walking-cycling-labels": {
                    "name": "Walking, cycling, etc., walking-cycling-labels",
                    "collapsed": false
                },
                "Walking, cycling, etc., surface": {
                    "name": "Walking, cycling, etc., surface",
                    "collapsed": false
                },
                "Transit, built": {"name": "Transit, built", "collapsed": false},
                "Land & water, water": {
                    "name": "Land & water, water",
                    "collapsed": false
                }
            },
            "mapbox:uiParadigm": "layers",
            "mapbox:decompiler": {
                "id": "'''+MapCode+'''",
                "componentVersion": "7.1.0",
                "strata": [
                    {
                        "id": "basic-v1",
                        "order": [
                            ["land-and-water", "land"],
                            ["land-and-water", "water"],
                            ["land-and-water", "built"],
                            ["transit", "built"],
                            ["buildings", "built"],
                            ["road-network", "tunnels-case"],
                            ["walking-cycling", "tunnels"],
                            ["road-network", "tunnels"],
                            ["transit", "ferries"],
                            ["walking-cycling", "surface"],
                            ["road-network", "surface"],
                            ["transit", "surface"],
                            ["road-network", "surface-icons"],
                            ["walking-cycling", "barriers-bridges"],
                            ["road-network", "bridges"],
                            ["transit", "bridges"],
                            ["road-network", "traffic-and-closures"],
                            ["buildings", "extruded"],
                            ["transit", "elevated"],
                            ["admin-boundaries", "admin"],
                            ["buildings", "building-labels"],
                            ["road-network", "road-labels"],
                            ["walking-cycling", "walking-cycling-labels"],
                            ["transit", "ferry-aerialway-labels"],
                            ["natural-features", "natural-labels"],
                            ["point-of-interest-labels", "poi-labels"],
                            ["transit", "transit-labels"],
                            ["place-labels", "place-labels"],
                            "'''+TileSet+'''"
                        ]
                    }
                ],
                "components": {
                    "road-network": "7.1.0",
                    "natural-features": "7.1.0",
                    "place-labels": "7.1.0",
                    "admin-boundaries": "7.1.0",
                    "point-of-interest-labels": "7.1.0",
                    "walking-cycling": "7.1.0",
                    "transit": "7.1.0",
                    "land-and-water": "7.1.0",
                    "buildings": "7.1.0"
                },
                "propConfig": {
                    "road-network": {
                        "color-base": "hsl(40, 48%, 92%)",
                        "color-road": "hsl(38, 55%, 100%)",
                        "roadNetwork": "Simple"
                    },
                    "natural-features": {
                        "color-base": "hsl(40, 48%, 92%)",
                        "color-water": "hsl(205, 76%, 70%)",
                        "color-poi": "hsl(26, 20%, 38%)"
                    },
                    "place-labels": {
                        "color-base": "hsl(40, 48%, 92%)",
                        "color-place-label": "hsl(0, 0%, 15%)",
                        "settlementSubdivisionsDensity": 3,
                        "settlementLabelStyle": "Text only"
                    },
                    "admin-boundaries": {
                        "color-base": "hsl(40, 48%, 92%)",
                        "color-place-label": "hsl(0, 0%, 15%)"
                    },
                    "point-of-interest-labels": {
                        "color-base": "hsl(40, 48%, 92%)",
                        "color-greenspace": "hsl(78, 50%, 73%)",
                        "color-greenspace-label": "hsl(76, 50%, 16%)",
                        "color-hospital": "hsl(3, 45%, 55%)",
                        "color-school": "hsl(40, 45%, 45%)",
                        "color-poi": "hsl(26, 20%, 38%)",
                        "density": 2
                    },
                    "walking-cycling": {
                        "color-base": "hsl(40, 48%, 92%)",
                        "color-road": "hsl(38, 55%, 100%)",
                        "color-greenspace": "hsl(78, 50%, 73%)",
                        "color-greenspace-label": "hsl(76, 50%, 16%)",
                        "walkingCyclingPisteBackground": false,
                        "golfHoleLabelLine": false,
                        "pedestrianPolygonFeatures": false
                    },
                    "transit": {
                        "color-airport": "hsl(225, 4%, 40%)",
                        "color-transit": "hsl(345, 6%, 40%)",
                        "aerialways": false,
                        "color-road": "hsl(38, 55%, 100%)",
                        "color-water": "hsl(205, 76%, 70%)",
                        "transitLabels": false,
                        "railways": false,
                        "ferries": false,
                        "color-base": "hsl(40, 48%, 92%)"
                    },
                    "land-and-water": {
                        "color-airport": "hsl(225, 4%, 40%)",
                        "color-hospital": "hsl(3, 45%, 55%)",
                        "landcover": false,
                        "color-greenspace": "hsl(78, 50%, 73%)",
                        "color-water": "hsl(205, 76%, 70%)",
                        "transitionLandOnZoom": false,
                        "waterStyle": "Simple",
                        "color-base": "hsl(40, 48%, 92%)",
                        "color-school": "hsl(40, 45%, 45%)"
                    },
                    "buildings": {
                        "color-base": "hsl(40, 48%, 92%)",
                        "houseNumbers": false
                    }
                }
            }
        },
        "center": ['''+CoordMapX+''', '''+CoordMapY+'''],
        "zoom": 11.65023001290617,
        "bearing": 0,
        "pitch": 0,
        "sources": {
            "composite": {
                "url": "mapbox://mapbox.mapbox-streets-v8,'''+URLtileset+'''",
                "type": "vector"
            }
        },
        "sprite": "mapbox://sprites/carmela-cucuzzella/cki95ew494au719ufyzwo9v7j/alroxy37js89g03iaz7wu12po",
        "glyphs": "mapbox://fonts/carmela-cucuzzella/{fontstack}/{range}.pbf",
        "layers": [
            {
                "id": "land",
                "type": "background",
                "metadata": {
                    "mapbox:featureComponent": "land-and-water",
                    "mapbox:group": "Land & water, land"
                },
                "layout": {},
                "paint": {"background-color": "hsl(40, 46%, 86%)"}
            },
            {
                "id": "national-park",
                "type": "fill",
                "metadata": {
                    "mapbox:featureComponent": "land-and-water",
                    "mapbox:group": "Land & water, land"
                },
                "source": "composite",
                "source-layer": "landuse_overlay",
                "minzoom": 5,
                "filter": ["==", ["get", "class"], "national_park"],
                "layout": {},
                "paint": {
                    "fill-color": "hsl(78, 50%, 73%)",
                    "fill-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        5,
                        0,
                        6,
                        0.5,
                        10,
                        0.5
                    ]
                }
            },
            {
                "id": "landuse",
                "type": "fill",
                "metadata": {
                    "mapbox:featureComponent": "land-and-water",
                    "mapbox:group": "Land & water, land"
                },
                "source": "composite",
                "source-layer": "landuse",
                "minzoom": 5,
                "filter": [
                    "match",
                    ["get", "class"],
                    ["park", "airport", "glacier", "pitch", "sand", "facility"],
                    true,
                    "cemetery",
                    true,
                    "school",
                    true,
                    "hospital",
                    true,
                    false
                ],
                "layout": {},
                "paint": {
                    "fill-color": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        15,
                        [
                            "match",
                            ["get", "class"],
                            "park",
                            "hsl(78, 50%, 73%)",
                            "airport",
                            "hsl(225, 49%, 88%)",
                            "cemetery",
                            "hsl(60, 49%, 79%)",
                            "glacier",
                            "hsl(205, 66%, 90%)",
                            "hospital",
                            "hsl(3, 47%, 84%)",
                            "pitch",
                            "hsl(78, 51%, 68%)",
                            "sand",
                            "hsl(43, 50%, 83%)",
                            "school",
                            "hsl(40, 47%, 78%)",
                            "hsl(40, 49%, 82%)"
                        ],
                        16,
                        [
                            "match",
                            ["get", "class"],
                            "park",
                            "hsl(78, 50%, 73%)",
                            "airport",
                            "hsl(225, 63%, 86%)",
                            "cemetery",
                            "hsl(60, 49%, 79%)",
                            "glacier",
                            "hsl(205, 66%, 90%)",
                            "hospital",
                            "hsl(3, 46%, 86%)",
                            "pitch",
                            "hsl(78, 51%, 68%)",
                            "sand",
                            "hsl(43, 50%, 83%)",
                            "school",
                            "hsl(40, 47%, 78%)",
                            "hsl(40, 49%, 82%)"
                        ]
                    ],
                    "fill-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        5,
                        0,
                        6,
                        ["match", ["get", "class"], "glacier", 0.5, 1]
                    ]
                }
            },
            {
                "id": "pitch-outline",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "land-and-water",
                    "mapbox:group": "Land & water, land"
                },
                "source": "composite",
                "source-layer": "landuse",
                "minzoom": 15,
                "filter": ["==", ["get", "class"], "pitch"],
                "layout": {},
                "paint": {"line-color": "hsl(60, 29%, 81%)"}
            },
            {
                "id": "waterway",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "land-and-water",
                    "mapbox:group": "Land & water, water"
                },
                "source": "composite",
                "source-layer": "waterway",
                "minzoom": 8,
                "layout": {
                    "line-cap": ["step", ["zoom"], "butt", 11, "round"],
                    "line-join": "round"
                },
                "paint": {
                    "line-color": "hsl(205, 76%, 70%)",
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.3],
                        ["zoom"],
                        9,
                        ["match", ["get", "class"], ["canal", "river"], 0.1, 0],
                        20,
                        ["match", ["get", "class"], ["canal", "river"], 8, 3]
                    ],
                    "line-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        8,
                        0,
                        8.5,
                        1
                    ]
                }
            },
            {
                "id": "water",
                "type": "fill",
                "metadata": {
                    "mapbox:featureComponent": "land-and-water",
                    "mapbox:group": "Land & water, water"
                },
                "source": "composite",
                "source-layer": "water",
                "layout": {},
                "paint": {"fill-color": "hsl(205, 76%, 70%)"}
            },
            {
                "id": "land-structure-polygon",
                "type": "fill",
                "metadata": {
                    "mapbox:featureComponent": "land-and-water",
                    "mapbox:group": "Land & water, built"
                },
                "source": "composite",
                "source-layer": "structure",
                "minzoom": 13,
                "filter": [
                    "all",
                    ["==", ["geometry-type"], "Polygon"],
                    ["==", ["get", "class"], "land"]
                ],
                "layout": {},
                "paint": {"fill-color": "hsl(40, 46%, 86%)"}
            },
            {
                "id": "land-structure-line",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "land-and-water",
                    "mapbox:group": "Land & water, built"
                },
                "source": "composite",
                "source-layer": "structure",
                "minzoom": 13,
                "filter": [
                    "all",
                    ["==", ["geometry-type"], "LineString"],
                    ["==", ["get", "class"], "land"]
                ],
                "layout": {"line-cap": "round"},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.99],
                        ["zoom"],
                        14,
                        0.75,
                        20,
                        40
                    ],
                    "line-color": "hsl(40, 46%, 86%)"
                }
            },
            {
                "id": "aeroway-polygon",
                "type": "fill",
                "metadata": {
                    "mapbox:featureComponent": "transit",
                    "mapbox:group": "Transit, built"
                },
                "source": "composite",
                "source-layer": "aeroway",
                "minzoom": 11,
                "filter": [
                    "all",
                    ["==", ["geometry-type"], "Polygon"],
                    [
                        "match",
                        ["get", "type"],
                        ["runway", "taxiway", "helipad"],
                        true,
                        false
                    ]
                ],
                "layout": {},
                "paint": {
                    "fill-color": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        15,
                        "hsl(225, 37%, 79%)",
                        16,
                        "hsl(225, 19%, 81%)"
                    ],
                    "fill-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        11,
                        0,
                        11.5,
                        1
                    ]
                }
            },
            {
                "id": "aeroway-line",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "transit",
                    "mapbox:group": "Transit, built"
                },
                "source": "composite",
                "source-layer": "aeroway",
                "minzoom": 9,
                "filter": ["==", ["geometry-type"], "LineString"],
                "layout": {},
                "paint": {
                    "line-color": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        15,
                        "hsl(225, 37%, 79%)",
                        16,
                        "hsl(225, 19%, 81%)"
                    ],
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        9,
                        ["match", ["get", "type"], "runway", 1, 0.5],
                        18,
                        ["match", ["get", "type"], "runway", 80, 20]
                    ]
                }
            },
            {
                "id": "building-outline",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "buildings",
                    "mapbox:group": "Buildings, built"
                },
                "source": "composite",
                "source-layer": "building",
                "minzoom": 15,
                "filter": [
                    "all",
                    ["!=", ["get", "type"], "building:part"],
                    ["==", ["get", "underground"], "false"]
                ],
                "layout": {},
                "paint": {
                    "line-color": "hsl(40, 42%, 77%)",
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        15,
                        0.75,
                        20,
                        3
                    ],
                    "line-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        15,
                        0,
                        16,
                        1
                    ]
                }
            },
            {
                "id": "building",
                "type": "fill",
                "metadata": {
                    "mapbox:featureComponent": "buildings",
                    "mapbox:group": "Buildings, built"
                },
                "source": "composite",
                "source-layer": "building",
                "minzoom": 15,
                "filter": [
                    "all",
                    ["!=", ["get", "type"], "building:part"],
                    ["==", ["get", "underground"], "false"]
                ],
                "layout": {},
                "paint": {
                    "fill-color": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        15,
                        "hsl(40, 45%, 83%)",
                        16,
                        "hsl(40, 41%, 82%)"
                    ],
                    "fill-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        15,
                        0,
                        16,
                        1
                    ],
                    "fill-outline-color": "hsl(40, 42%, 77%)"
                }
            },
            {
                "id": "tunnel-path",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., tunnels"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 14,
                "filter": [
                    "all",
                    ["==", ["get", "structure"], "tunnel"],
                    ["==", ["get", "class"], "path"],
                    ["!=", ["get", "type"], "steps"],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        15,
                        1,
                        18,
                        4
                    ],
                    "line-color": "hsl(40, 42%, 77%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [1, 0]],
                        15,
                        ["literal", [1.75, 1]],
                        16,
                        ["literal", [1, 0.75]],
                        17,
                        ["literal", [1, 0.5]]
                    ]
                }
            },
            {
                "id": "tunnel-steps",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., tunnels"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 14,
                "filter": [
                    "all",
                    ["==", ["get", "structure"], "tunnel"],
                    ["==", ["get", "type"], "steps"],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        15,
                        1,
                        16,
                        1.6,
                        18,
                        6
                    ],
                    "line-color": "hsl(40, 42%, 77%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [1, 0]],
                        15,
                        ["literal", [1.75, 1]],
                        16,
                        ["literal", [1, 0.75]],
                        17,
                        ["literal", [0.3, 0.3]]
                    ]
                }
            },
            {
                "id": "tunnel-pedestrian",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., tunnels"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 13,
                "filter": [
                    "all",
                    ["==", ["get", "structure"], "tunnel"],
                    ["==", ["get", "class"], "pedestrian"],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        14,
                        0.5,
                        18,
                        12
                    ],
                    "line-color": "hsl(40, 42%, 77%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [1, 0]],
                        15,
                        ["literal", [1.5, 0.4]],
                        16,
                        ["literal", [1, 0.2]]
                    ]
                }
            },
            {
                "id": "tunnel-simple",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "road-network",
                    "mapbox:group": "Road network, tunnels"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 13,
                "filter": [
                    "all",
                    ["==", ["get", "structure"], "tunnel"],
                    [
                        "step",
                        ["zoom"],
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "secondary",
                                "tertiary",
                                "street",
                                "street_limited",
                                "primary_link",
                                "track"
                            ],
                            true,
                            false
                        ],
                        14,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "primary_link",
                                "secondary",
                                "secondary_link",
                                "tertiary",
                                "tertiary_link",
                                "street",
                                "street_limited",
                                "service",
                                "track"
                            ],
                            true,
                            false
                        ]
                    ],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        13,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            4,
                            ["secondary", "tertiary"],
                            2.5,
                            [
                                "motorway_link",
                                "trunk_link",
                                "street",
                                "street_limited",
                                "primary_link"
                            ],
                            1,
                            0.5
                        ],
                        18,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            32,
                            ["secondary", "tertiary"],
                            26,
                            [
                                "motorway_link",
                                "trunk_link",
                                "street",
                                "street_limited",
                                "primary_link"
                            ],
                            18,
                            12
                        ]
                    ],
                    "line-color": "hsl(38, 55%, 93%)"
                }
            },
            {
                "id": "road-path",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., surface"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 12,
                "filter": [
                    "all",
                    ["==", ["get", "class"], "path"],
                    [
                        "step",
                        ["zoom"],
                        [
                            "!",
                            [
                                "match",
                                ["get", "type"],
                                ["steps", "sidewalk", "crossing"],
                                true,
                                false
                            ]
                        ],
                        16,
                        ["!=", ["get", "type"], "steps"]
                    ],
                    ["match", ["get", "structure"], ["none", "ford"], true, false],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {"line-join": ["step", ["zoom"], "miter", 14, "round"]},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        13,
                        0.5,
                        14,
                        1,
                        15,
                        1,
                        18,
                        4
                    ],
                    "line-color": "hsl(40, 47%, 96%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [4, 0.3]],
                        15,
                        ["literal", [1.75, 0.3]],
                        16,
                        ["literal", [1, 0.3]],
                        17,
                        ["literal", [1, 0.25]]
                    ]
                }
            },
            {
                "id": "road-steps",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., surface"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 14,
                "filter": [
                    "all",
                    ["==", ["get", "type"], "steps"],
                    ["match", ["get", "structure"], ["none", "ford"], true, false],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {"line-join": "round"},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        15,
                        1,
                        16,
                        1.6,
                        18,
                        6
                    ],
                    "line-color": "hsl(40, 47%, 96%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [1, 0]],
                        15,
                        ["literal", [1.75, 1]],
                        16,
                        ["literal", [1, 0.75]],
                        17,
                        ["literal", [0.3, 0.3]]
                    ]
                }
            },
            {
                "id": "road-pedestrian",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., surface"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 12,
                "filter": [
                    "all",
                    ["==", ["get", "class"], "pedestrian"],
                    ["match", ["get", "structure"], ["none", "ford"], true, false],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {"line-join": ["step", ["zoom"], "miter", 14, "round"]},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        14,
                        0.5,
                        18,
                        12
                    ],
                    "line-color": "hsl(40, 47%, 96%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [1, 0]],
                        15,
                        ["literal", [1.5, 0.4]],
                        16,
                        ["literal", [1, 0.2]]
                    ]
                }
            },
            {
                "id": "road-simple",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "road-network",
                    "mapbox:group": "Road network, surface"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 5,
                "filter": [
                    "all",
                    [
                        "step",
                        ["zoom"],
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk"],
                            true,
                            false
                        ],
                        6,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            true,
                            false
                        ],
                        8,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary", "secondary"],
                            true,
                            false
                        ],
                        10,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "trunk",
                                "primary",
                                "secondary",
                                "tertiary",
                                "motorway_link",
                                "trunk_link"
                            ],
                            true,
                            false
                        ],
                        11,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "secondary",
                                "tertiary",
                                "street"
                            ],
                            true,
                            false
                        ],
                        12,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "secondary",
                                "tertiary",
                                "street",
                                "street_limited",
                                "primary_link"
                            ],
                            true,
                            false
                        ],
                        13,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "secondary",
                                "tertiary",
                                "street",
                                "street_limited",
                                "primary_link",
                                "track"
                            ],
                            true,
                            false
                        ],
                        14,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "primary_link",
                                "secondary",
                                "secondary_link",
                                "tertiary",
                                "tertiary_link",
                                "street",
                                "street_limited",
                                "service",
                                "track"
                            ],
                            true,
                            false
                        ]
                    ],
                    ["match", ["get", "structure"], ["none", "ford"], true, false],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {
                    "line-cap": ["step", ["zoom"], "butt", 14, "round"],
                    "line-join": ["step", ["zoom"], "miter", 14, "round"]
                },
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        5,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            0.75,
                            ["secondary", "tertiary"],
                            0.1,
                            0
                        ],
                        13,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            4,
                            ["secondary", "tertiary"],
                            2.5,
                            [
                                "motorway_link",
                                "trunk_link",
                                "primary_link",
                                "street",
                                "street_limited"
                            ],
                            1,
                            0.5
                        ],
                        18,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            32,
                            ["secondary", "tertiary"],
                            26,
                            [
                                "motorway_link",
                                "trunk_link",
                                "primary_link",
                                "street",
                                "street_limited"
                            ],
                            18,
                            10
                        ]
                    ],
                    "line-color": [
                        "match",
                        ["get", "class"],
                        [
                            "primary_link",
                            "secondary_link",
                            "tertiary_link",
                            "street",
                            "street_limited",
                            "service",
                            "track"
                        ],
                        "hsl(38, 55%, 95%)",
                        "hsl(38, 55%, 100%)"
                    ]
                }
            },
            {
                "id": "bridge-path",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., barriers-bridges"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 14,
                "filter": [
                    "all",
                    ["==", ["get", "structure"], "bridge"],
                    ["==", ["get", "class"], "path"],
                    ["==", ["geometry-type"], "LineString"],
                    ["!=", ["get", "type"], "steps"]
                ],
                "layout": {"line-join": "round"},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        15,
                        1,
                        18,
                        4
                    ],
                    "line-color": "hsl(40, 47%, 96%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [4, 0.3]],
                        15,
                        ["literal", [1.75, 0.3]],
                        16,
                        ["literal", [1, 0.3]],
                        17,
                        ["literal", [1, 0.25]]
                    ]
                }
            },
            {
                "id": "bridge-steps",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., barriers-bridges"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 14,
                "filter": [
                    "all",
                    ["==", ["get", "type"], "steps"],
                    ["==", ["get", "structure"], "bridge"],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {"line-join": "round"},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        15,
                        1,
                        16,
                        1.6,
                        18,
                        6
                    ],
                    "line-color": "hsl(40, 47%, 96%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [1, 0]],
                        15,
                        ["literal", [1.75, 1]],
                        16,
                        ["literal", [1, 0.75]],
                        17,
                        ["literal", [0.3, 0.3]]
                    ]
                }
            },
            {
                "id": "bridge-pedestrian",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., barriers-bridges"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 13,
                "filter": [
                    "all",
                    ["==", ["get", "structure"], "bridge"],
                    ["==", ["get", "class"], "pedestrian"],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {"line-join": "round"},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        14,
                        0.5,
                        18,
                        12
                    ],
                    "line-color": "hsl(40, 47%, 96%)",
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [1, 0]],
                        15,
                        ["literal", [1.5, 0.4]],
                        16,
                        ["literal", [1, 0.2]]
                    ]
                }
            },
            {
                "id": "bridge-case-simple",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "road-network",
                    "mapbox:group": "Road network, bridges"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 13,
                "filter": [
                    "all",
                    ["==", ["get", "structure"], "bridge"],
                    [
                        "step",
                        ["zoom"],
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "secondary",
                                "tertiary",
                                "street",
                                "street_limited",
                                "primary_link",
                                "track"
                            ],
                            true,
                            false
                        ],
                        14,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "primary_link",
                                "secondary",
                                "secondary_link",
                                "tertiary",
                                "tertiary_link",
                                "street",
                                "street_limited",
                                "service",
                                "track"
                            ],
                            true,
                            false
                        ]
                    ],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {"line-join": ["step", ["zoom"], "miter", 14, "round"]},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        13,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            6,
                            ["secondary", "tertiary"],
                            4,
                            [
                                "motorway_link",
                                "trunk_link",
                                "street",
                                "street_limited",
                                "primary_link"
                            ],
                            2.5,
                            1.25
                        ],
                        18,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            36,
                            ["secondary", "tertiary"],
                            30,
                            [
                                "motorway_link",
                                "trunk_link",
                                "street",
                                "street_limited",
                                "primary_link"
                            ],
                            22,
                            16
                        ]
                    ],
                    "line-color": "hsl(40, 46%, 86%)"
                }
            },
            {
                "id": "bridge-simple",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "road-network",
                    "mapbox:group": "Road network, bridges"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 13,
                "filter": [
                    "all",
                    ["==", ["get", "structure"], "bridge"],
                    [
                        "step",
                        ["zoom"],
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk"],
                            true,
                            false
                        ],
                        13,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "secondary",
                                "tertiary",
                                "street",
                                "street_limited",
                                "primary_link",
                                "track"
                            ],
                            true,
                            false
                        ],
                        14,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "motorway_link",
                                "trunk",
                                "trunk_link",
                                "primary",
                                "primary_link",
                                "secondary",
                                "secondary_link",
                                "tertiary",
                                "tertiary_link",
                                "street",
                                "street_limited",
                                "service",
                                "track"
                            ],
                            true,
                            false
                        ]
                    ],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {
                    "line-cap": ["step", ["zoom"], "butt", 14, "round"],
                    "line-join": ["step", ["zoom"], "miter", 14, "round"]
                },
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["exponential", 1.5],
                        ["zoom"],
                        13,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            4,
                            ["secondary", "tertiary"],
                            2.5,
                            [
                                "motorway_link",
                                "trunk_link",
                                "street",
                                "street_limited",
                                "primary_link"
                            ],
                            1,
                            0.5
                        ],
                        18,
                        [
                            "match",
                            ["get", "class"],
                            ["motorway", "trunk", "primary"],
                            32,
                            ["secondary", "tertiary"],
                            26,
                            [
                                "motorway_link",
                                "trunk_link",
                                "street",
                                "street_limited",
                                "primary_link"
                            ],
                            18,
                            12
                        ]
                    ],
                    "line-color": [
                        "match",
                        ["get", "class"],
                        [
                            "primary_link",
                            "secondary_link",
                            "tertiary_link",
                            "street",
                            "street_limited",
                            "service",
                            "track"
                        ],
                        "hsl(38, 55%, 95%)",
                        "hsl(38, 55%, 100%)"
                    ]
                }
            },
            {
                "id": "admin-1-boundary-bg",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "admin-boundaries",
                    "mapbox:group": "Administrative boundaries, admin"
                },
                "source": "composite",
                "source-layer": "admin",
                "minzoom": 7,
                "filter": [
                    "all",
                    ["==", ["get", "admin_level"], 1],
                    ["==", ["get", "maritime"], "false"],
                    ["match", ["get", "worldview"], ["all", "US"], true, false]
                ],
                "layout": {"line-join": "bevel"},
                "paint": {
                    "line-color": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        8,
                        "hsl(40, 46%, 86%)",
                        16,
                        "hsl(0, 0%, 87%)"
                    ],
                    "line-width": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        7,
                        3.75,
                        12,
                        5.5
                    ],
                    "line-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        7,
                        0,
                        8,
                        0.75
                    ],
                    "line-dasharray": [1, 0],
                    "line-blur": ["interpolate", ["linear"], ["zoom"], 3, 0, 8, 3]
                }
            },
            {
                "id": "admin-0-boundary-bg",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "admin-boundaries",
                    "mapbox:group": "Administrative boundaries, admin"
                },
                "source": "composite",
                "source-layer": "admin",
                "minzoom": 1,
                "filter": [
                    "all",
                    ["==", ["get", "admin_level"], 0],
                    ["==", ["get", "maritime"], "false"],
                    ["match", ["get", "worldview"], ["all", "US"], true, false]
                ],
                "layout": {},
                "paint": {
                    "line-width": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        3,
                        3.5,
                        10,
                        8
                    ],
                    "line-color": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        6,
                        "hsl(40, 46%, 86%)",
                        8,
                        "hsl(0, 0%, 87%)"
                    ],
                    "line-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        3,
                        0,
                        4,
                        0.5
                    ],
                    "line-blur": ["interpolate", ["linear"], ["zoom"], 3, 0, 10, 2]
                }
            },
            {
                "id": "admin-1-boundary",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "admin-boundaries",
                    "mapbox:group": "Administrative boundaries, admin"
                },
                "source": "composite",
                "source-layer": "admin",
                "minzoom": 2,
                "filter": [
                    "all",
                    ["==", ["get", "admin_level"], 1],
                    ["==", ["get", "maritime"], "false"],
                    ["match", ["get", "worldview"], ["all", "US"], true, false]
                ],
                "layout": {"line-join": "round", "line-cap": "round"},
                "paint": {
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [2, 0]],
                        7,
                        ["literal", [2, 2, 6, 2]]
                    ],
                    "line-width": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        7,
                        0.75,
                        12,
                        1.5
                    ],
                    "line-opacity": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        2,
                        0,
                        3,
                        1
                    ],
                    "line-color": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        3,
                        "hsl(0, 0%, 77%)",
                        7,
                        "hsl(0, 0%, 62%)"
                    ]
                }
            },
            {
                "id": "admin-0-boundary",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "admin-boundaries",
                    "mapbox:group": "Administrative boundaries, admin"
                },
                "source": "composite",
                "source-layer": "admin",
                "minzoom": 1,
                "filter": [
                    "all",
                    ["==", ["get", "admin_level"], 0],
                    ["==", ["get", "disputed"], "false"],
                    ["==", ["get", "maritime"], "false"],
                    ["match", ["get", "worldview"], ["all", "US"], true, false]
                ],
                "layout": {"line-join": "round", "line-cap": "round"},
                "paint": {
                    "line-color": "hsl(0, 0%, 51%)",
                    "line-width": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        3,
                        0.5,
                        10,
                        2
                    ],
                    "line-dasharray": [10, 0]
                }
            },
            {
                "id": "admin-0-boundary-disputed",
                "type": "line",
                "metadata": {
                    "mapbox:featureComponent": "admin-boundaries",
                    "mapbox:group": "Administrative boundaries, admin"
                },
                "source": "composite",
                "source-layer": "admin",
                "minzoom": 1,
                "filter": [
                    "all",
                    ["==", ["get", "disputed"], "true"],
                    ["==", ["get", "admin_level"], 0],
                    ["==", ["get", "maritime"], "false"],
                    ["match", ["get", "worldview"], ["all", "US"], true, false]
                ],
                "layout": {"line-join": "round"},
                "paint": {
                    "line-color": "hsl(0, 0%, 51%)",
                    "line-width": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        3,
                        0.5,
                        10,
                        2
                    ],
                    "line-dasharray": [
                        "step",
                        ["zoom"],
                        ["literal", [3.25, 3.25]],
                        6,
                        ["literal", [2.5, 2.5]],
                        7,
                        ["literal", [2, 2.25]],
                        8,
                        ["literal", [1.75, 2]]
                    ]
                }
            },
            {
                "id": "road-label-simple",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "road-network",
                    "mapbox:group": "Road network, road-labels"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 12,
                "filter": [
                    "match",
                    ["get", "class"],
                    [
                        "motorway",
                        "trunk",
                        "primary",
                        "secondary",
                        "tertiary",
                        "street",
                        "street_limited"
                    ],
                    true,
                    false
                ],
                "layout": {
                    "text-size": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        10,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "trunk",
                                "primary",
                                "secondary",
                                "tertiary"
                            ],
                            10,
                            9
                        ],
                        18,
                        [
                            "match",
                            ["get", "class"],
                            [
                                "motorway",
                                "trunk",
                                "primary",
                                "secondary",
                                "tertiary"
                            ],
                            16,
                            14
                        ]
                    ],
                    "text-max-angle": 30,
                    "text-font": ["DIN Pro Regular", "Arial Unicode MS Regular"],
                    "symbol-placement": "line",
                    "text-padding": 1,
                    "text-rotation-alignment": "map",
                    "text-pitch-alignment": "viewport",
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]],
                    "text-letter-spacing": 0.01
                },
                "paint": {
                    "text-color": "hsl(40, 47%, 41%)",
                    "text-halo-color": "hsl(38, 55%, 100%)",
                    "text-halo-width": 1
                }
            },
            {
                "id": "path-pedestrian-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "walking-cycling",
                    "mapbox:group": "Walking, cycling, etc., walking-cycling-labels"
                },
                "source": "composite",
                "source-layer": "road",
                "minzoom": 12,
                "filter": [
                    "step",
                    ["zoom"],
                    ["match", ["get", "class"], ["pedestrian"], true, false],
                    15,
                    ["match", ["get", "class"], ["path", "pedestrian"], true, false]
                ],
                "layout": {
                    "text-size": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        10,
                        ["match", ["get", "class"], "pedestrian", 9, 6.5],
                        18,
                        ["match", ["get", "class"], "pedestrian", 14, 13]
                    ],
                    "text-max-angle": 30,
                    "text-font": ["DIN Pro Regular", "Arial Unicode MS Regular"],
                    "symbol-placement": "line",
                    "text-padding": 1,
                    "text-rotation-alignment": "map",
                    "text-pitch-alignment": "viewport",
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]],
                    "text-letter-spacing": 0.01
                },
                "paint": {
                    "text-color": "hsl(40, 47%, 41%)",
                    "text-halo-color": "hsl(40, 46%, 95%)",
                    "text-halo-width": 1,
                    "text-halo-blur": 1
                }
            },
            {
                "id": "waterway-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "natural-features",
                    "mapbox:group": "Natural features, natural-labels"
                },
                "source": "composite",
                "source-layer": "natural_label",
                "minzoom": 13,
                "filter": [
                    "all",
                    [
                        "match",
                        ["get", "class"],
                        ["canal", "river", "stream"],
                        ["match", ["get", "worldview"], ["all", "US"], true, false],
                        ["disputed_canal", "disputed_river", "disputed_stream"],
                        [
                            "all",
                            ["==", ["get", "disputed"], "true"],
                            [
                                "match",
                                ["get", "worldview"],
                                ["all", "US"],
                                true,
                                false
                            ]
                        ],
                        false
                    ],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {
                    "text-font": ["DIN Pro Italic", "Arial Unicode MS Regular"],
                    "text-max-angle": 30,
                    "symbol-spacing": [
                        "interpolate",
                        ["linear", 1],
                        ["zoom"],
                        15,
                        250,
                        17,
                        400
                    ],
                    "text-size": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        13,
                        12,
                        18,
                        16
                    ],
                    "symbol-placement": "line",
                    "text-pitch-alignment": "viewport",
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]]
                },
                "paint": {"text-color": "hsl(205, 44%, 90%)"}
            },
            {
                "id": "natural-line-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "natural-features",
                    "mapbox:group": "Natural features, natural-labels"
                },
                "source": "composite",
                "source-layer": "natural_label",
                "minzoom": 4,
                "filter": [
                    "all",
                    [
                        "match",
                        ["get", "class"],
                        ["glacier", "landform"],
                        ["match", ["get", "worldview"], ["all", "US"], true, false],
                        ["disputed_glacier", "disputed_landform"],
                        [
                            "all",
                            ["==", ["get", "disputed"], "true"],
                            [
                                "match",
                                ["get", "worldview"],
                                ["all", "US"],
                                true,
                                false
                            ]
                        ],
                        false
                    ],
                    ["==", ["geometry-type"], "LineString"],
                    ["<=", ["get", "filterrank"], 2]
                ],
                "layout": {
                    "text-size": [
                        "step",
                        ["zoom"],
                        ["step", ["get", "sizerank"], 18, 5, 12],
                        17,
                        ["step", ["get", "sizerank"], 18, 13, 12]
                    ],
                    "text-max-angle": 30,
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]],
                    "text-font": ["DIN Pro Medium", "Arial Unicode MS Regular"],
                    "symbol-placement": "line-center",
                    "text-pitch-alignment": "viewport"
                },
                "paint": {
                    "text-halo-width": 0.5,
                    "text-halo-color": "hsl(40, 53%, 100%)",
                    "text-halo-blur": 0.5,
                    "text-color": [
                        "step",
                        ["zoom"],
                        [
                            "step",
                            ["get", "sizerank"],
                            "hsl(26, 15%, 48%)",
                            5,
                            "hsl(26, 20%, 38%)"
                        ],
                        17,
                        [
                            "step",
                            ["get", "sizerank"],
                            "hsl(26, 15%, 48%)",
                            13,
                            "hsl(26, 20%, 38%)"
                        ]
                    ]
                }
            },
            {
                "id": "natural-point-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "natural-features",
                    "mapbox:group": "Natural features, natural-labels"
                },
                "source": "composite",
                "source-layer": "natural_label",
                "minzoom": 4,
                "filter": [
                    "all",
                    [
                        "match",
                        ["get", "class"],
                        ["dock", "glacier", "landform", "water_feature", "wetland"],
                        ["match", ["get", "worldview"], ["all", "US"], true, false],
                        [
                            "disputed_dock",
                            "disputed_glacier",
                            "disputed_landform",
                            "disputed_water_feature",
                            "disputed_wetland"
                        ],
                        [
                            "all",
                            ["==", ["get", "disputed"], "true"],
                            [
                                "match",
                                ["get", "worldview"],
                                ["all", "US"],
                                true,
                                false
                            ]
                        ],
                        false
                    ],
                    ["==", ["geometry-type"], "Point"],
                    ["<=", ["get", "filterrank"], 2]
                ],
                "layout": {
                    "text-size": [
                        "step",
                        ["zoom"],
                        ["step", ["get", "sizerank"], 18, 5, 12],
                        17,
                        ["step", ["get", "sizerank"], 18, 13, 12]
                    ],
                    "icon-image": [
                        "step",
                        ["zoom"],
                        ["concat", ["get", "maki"], "-11"],
                        15,
                        ["concat", ["get", "maki"], "-15"]
                    ],
                    "text-font": ["DIN Pro Medium", "Arial Unicode MS Regular"],
                    "text-offset": [
                        "step",
                        ["zoom"],
                        [
                            "step",
                            ["get", "sizerank"],
                            ["literal", [0, 0]],
                            5,
                            ["literal", [0, 0.75]]
                        ],
                        17,
                        [
                            "step",
                            ["get", "sizerank"],
                            ["literal", [0, 0]],
                            13,
                            ["literal", [0, 0.75]]
                        ]
                    ],
                    "text-anchor": [
                        "step",
                        ["zoom"],
                        ["step", ["get", "sizerank"], "center", 5, "top"],
                        17,
                        ["step", ["get", "sizerank"], "center", 13, "top"]
                    ],
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]]
                },
                "paint": {
                    "icon-opacity": [
                        "step",
                        ["zoom"],
                        ["step", ["get", "sizerank"], 0, 5, 1],
                        17,
                        ["step", ["get", "sizerank"], 0, 13, 1]
                    ],
                    "text-halo-color": "hsl(40, 53%, 100%)",
                    "text-halo-width": 0.5,
                    "text-halo-blur": 0.5,
                    "text-color": [
                        "step",
                        ["zoom"],
                        [
                            "step",
                            ["get", "sizerank"],
                            "hsl(26, 15%, 48%)",
                            5,
                            "hsl(26, 20%, 38%)"
                        ],
                        17,
                        [
                            "step",
                            ["get", "sizerank"],
                            "hsl(26, 15%, 48%)",
                            13,
                            "hsl(26, 20%, 38%)"
                        ]
                    ]
                }
            },
            {
                "id": "water-line-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "natural-features",
                    "mapbox:group": "Natural features, natural-labels"
                },
                "source": "composite",
                "source-layer": "natural_label",
                "filter": [
                    "all",
                    [
                        "match",
                        ["get", "class"],
                        ["bay", "ocean", "reservoir", "sea", "water"],
                        ["match", ["get", "worldview"], ["all", "US"], true, false],
                        [
                            "disputed_bay",
                            "disputed_ocean",
                            "disputed_reservoir",
                            "disputed_sea",
                            "disputed_water"
                        ],
                        [
                            "all",
                            ["==", ["get", "disputed"], "true"],
                            [
                                "match",
                                ["get", "worldview"],
                                ["all", "US"],
                                true,
                                false
                            ]
                        ],
                        false
                    ],
                    ["==", ["geometry-type"], "LineString"]
                ],
                "layout": {
                    "text-size": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        7,
                        ["step", ["get", "sizerank"], 20, 6, 18, 12, 12],
                        10,
                        ["step", ["get", "sizerank"], 15, 9, 12],
                        18,
                        ["step", ["get", "sizerank"], 15, 9, 14]
                    ],
                    "text-max-angle": 30,
                    "text-letter-spacing": [
                        "match",
                        ["get", "class"],
                        "ocean",
                        0.25,
                        ["sea", "bay"],
                        0.15,
                        0
                    ],
                    "text-font": ["DIN Pro Italic", "Arial Unicode MS Regular"],
                    "symbol-placement": "line-center",
                    "text-pitch-alignment": "viewport",
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]]
                },
                "paint": {
                    "text-color": [
                        "match",
                        ["get", "class"],
                        ["bay", "ocean", "sea"],
                        "hsl(205, 72%, 90%)",
                        "hsl(205, 44%, 90%)"
                    ]
                }
            },
            {
                "id": "water-point-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "natural-features",
                    "mapbox:group": "Natural features, natural-labels"
                },
                "source": "composite",
                "source-layer": "natural_label",
                "filter": [
                    "all",
                    [
                        "match",
                        ["get", "class"],
                        ["bay", "ocean", "reservoir", "sea", "water"],
                        ["match", ["get", "worldview"], ["all", "US"], true, false],
                        [
                            "disputed_bay",
                            "disputed_ocean",
                            "disputed_reservoir",
                            "disputed_sea",
                            "disputed_water"
                        ],
                        [
                            "all",
                            ["==", ["get", "disputed"], "true"],
                            [
                                "match",
                                ["get", "worldview"],
                                ["all", "US"],
                                true,
                                false
                            ]
                        ],
                        false
                    ],
                    ["==", ["geometry-type"], "Point"]
                ],
                "layout": {
                    "text-line-height": 1.3,
                    "text-size": [
                        "interpolate",
                        ["linear"],
                        ["zoom"],
                        7,
                        ["step", ["get", "sizerank"], 20, 6, 15, 12, 12],
                        10,
                        ["step", ["get", "sizerank"], 15, 9, 12]
                    ],
                    "text-font": ["DIN Pro Italic", "Arial Unicode MS Regular"],
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]],
                    "text-letter-spacing": [
                        "match",
                        ["get", "class"],
                        "ocean",
                        0.25,
                        ["bay", "sea"],
                        0.15,
                        0.01
                    ],
                    "text-max-width": [
                        "match",
                        ["get", "class"],
                        "ocean",
                        4,
                        "sea",
                        5,
                        ["bay", "water"],
                        7,
                        10
                    ]
                },
                "paint": {
                    "text-color": [
                        "match",
                        ["get", "class"],
                        ["bay", "ocean", "sea"],
                        "hsl(205, 72%, 90%)",
                        "hsl(205, 44%, 90%)"
                    ]
                }
            },
            {
                "id": "poi-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "point-of-interest-labels",
                    "mapbox:group": "Point of interest labels, poi-labels"
                },
                "source": "composite",
                "source-layer": "poi_label",
                "minzoom": 6,
                "filter": [
                    "<=",
                    ["get", "filterrank"],
                    ["+", ["step", ["zoom"], 0, 16, 1, 17, 2], 2]
                ],
                "layout": {
                    "text-size": [
                        "step",
                        ["zoom"],
                        ["step", ["get", "sizerank"], 18, 5, 12],
                        17,
                        ["step", ["get", "sizerank"], 18, 13, 12]
                    ],
                    "icon-image": [
                        "step",
                        ["zoom"],
                        [
                            "case",
                            ["has", "maki_beta"],
                            ["image", ["concat", ["get", "maki_beta"], "-11"]],
                            ["image", ["concat", ["get", "maki"], "-11"]]
                        ],
                        15,
                        [
                            "case",
                            ["has", "maki_beta"],
                            ["image", ["concat", ["get", "maki_beta"], "-15"]],
                            ["image", ["concat", ["get", "maki"], "-15"]]
                        ]
                    ],
                    "text-font": ["DIN Pro Medium", "Arial Unicode MS Regular"],
                    "text-offset": [
                        "step",
                        ["zoom"],
                        [
                            "step",
                            ["get", "sizerank"],
                            ["literal", [0, 0]],
                            5,
                            ["literal", [0, 0.75]]
                        ],
                        17,
                        [
                            "step",
                            ["get", "sizerank"],
                            ["literal", [0, 0]],
                            13,
                            ["literal", [0, 0.75]]
                        ]
                    ],
                    "text-anchor": [
                        "step",
                        ["zoom"],
                        ["step", ["get", "sizerank"], "center", 5, "top"],
                        17,
                        ["step", ["get", "sizerank"], "center", 13, "top"]
                    ],
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]]
                },
                "paint": {
                    "icon-opacity": [
                        "step",
                        ["zoom"],
                        ["step", ["get", "sizerank"], 0, 5, 1],
                        17,
                        ["step", ["get", "sizerank"], 0, 13, 1]
                    ],
                    "text-halo-color": [
                        "match",
                        ["get", "class"],
                        "park_like",
                        "hsl(78, 55%, 100%)",
                        "education",
                        "hsl(40, 52%, 100%)",
                        "medical",
                        "hsl(3, 51%, 100%)",
                        "hsl(40, 53%, 100%)"
                    ],
                    "text-halo-width": 0.5,
                    "text-halo-blur": 0.5,
                    "text-color": [
                        "step",
                        ["zoom"],
                        [
                            "step",
                            ["get", "sizerank"],
                            [
                                "match",
                                ["get", "class"],
                                "food_and_drink",
                                "hsl(22, 44%, 61%)",
                                "park_like",
                                "hsl(76, 51%, 26%)",
                                "education",
                                "hsl(40, 18%, 45%)",
                                "medical",
                                "hsl(3, 18%, 55%)",
                                "hsl(26, 15%, 48%)"
                            ],
                            5,
                            [
                                "match",
                                ["get", "class"],
                                "food_and_drink",
                                "hsl(22, 68%, 44%)",
                                "park_like",
                                "hsl(76, 50%, 15%)",
                                "education",
                                "hsl(40, 45%, 25%)",
                                "medical",
                                "hsl(3, 24%, 45%)",
                                "hsl(26, 20%, 38%)"
                            ]
                        ],
                        17,
                        [
                            "step",
                            ["get", "sizerank"],
                            [
                                "match",
                                ["get", "class"],
                                "food_and_drink",
                                "hsl(22, 44%, 61%)",
                                "park_like",
                                "hsl(76, 51%, 26%)",
                                "education",
                                "hsl(40, 18%, 45%)",
                                "medical",
                                "hsl(3, 18%, 55%)",
                                "hsl(26, 15%, 48%)"
                            ],
                            13,
                            [
                                "match",
                                ["get", "class"],
                                "food_and_drink",
                                "hsl(22, 68%, 44%)",
                                "park_like",
                                "hsl(76, 50%, 15%)",
                                "education",
                                "hsl(40, 45%, 25%)",
                                "medical",
                                "hsl(3, 24%, 45%)",
                                "hsl(26, 20%, 38%)"
                            ]
                        ]
                    ]
                }
            },
            {
                "id": "airport-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "transit",
                    "mapbox:group": "Transit, transit-labels"
                },
                "source": "composite",
                "source-layer": "airport_label",
                "minzoom": 8,
                "filter": [
                    "match",
                    ["get", "class"],
                    ["military", "civil"],
                    ["match", ["get", "worldview"], ["all", "US"], true, false],
                    ["disputed_military", "disputed_civil"],
                    [
                        "all",
                        ["==", ["get", "disputed"], "true"],
                        ["match", ["get", "worldview"], ["all", "US"], true, false]
                    ],
                    false
                ],
                "layout": {
                    "text-line-height": 1.1,
                    "text-size": ["step", ["get", "sizerank"], 18, 9, 12],
                    "icon-image": [
                        "step",
                        ["get", "sizerank"],
                        ["concat", ["get", "maki"], "-15"],
                        9,
                        ["concat", ["get", "maki"], "-11"]
                    ],
                    "text-font": ["DIN Pro Medium", "Arial Unicode MS Regular"],
                    "text-offset": [0, 0.75],
                    "text-rotation-alignment": "viewport",
                    "text-anchor": "top",
                    "text-field": [
                        "step",
                        ["get", "sizerank"],
                        ["coalesce", ["get", "name_en"], ["get", "name"]],
                        15,
                        ["get", "ref"]
                    ],
                    "text-letter-spacing": 0.01,
                    "text-max-width": 9
                },
                "paint": {
                    "text-color": "hsl(225, 4%, 40%)",
                    "text-halo-color": "hsl(225, 68%, 100%)",
                    "text-halo-width": 1
                }
            },
            {
                "id": "settlement-subdivision-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "place-labels",
                    "mapbox:group": "Place labels, place-labels"
                },
                "source": "composite",
                "source-layer": "place_label",
                "minzoom": 10,
                "maxzoom": 15,
                "filter": [
                    "all",
                    [
                        "match",
                        ["get", "class"],
                        "settlement_subdivision",
                        ["match", ["get", "worldview"], ["all", "US"], true, false],
                        "disputed_settlement_subdivision",
                        [
                            "all",
                            ["==", ["get", "disputed"], "true"],
                            [
                                "match",
                                ["get", "worldview"],
                                ["all", "US"],
                                true,
                                false
                            ]
                        ],
                        false
                    ],
                    ["<=", ["get", "filterrank"], 3]
                ],
                "layout": {
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]],
                    "text-transform": "uppercase",
                    "text-font": ["DIN Pro Regular", "Arial Unicode MS Regular"],
                    "text-letter-spacing": [
                        "match",
                        ["get", "type"],
                        "suburb",
                        0.15,
                        0.1
                    ],
                    "text-max-width": 7,
                    "text-padding": 3,
                    "text-size": [
                        "interpolate",
                        ["cubic-bezier", 0.5, 0, 1, 1],
                        ["zoom"],
                        11,
                        ["match", ["get", "type"], "suburb", 11, 10.5],
                        15,
                        ["match", ["get", "type"], "suburb", 15, 14]
                    ]
                },
                "paint": {
                    "text-halo-color": "hsla(40, 53%, 100%, 0.75)",
                    "text-halo-width": 1,
                    "text-color": "hsl(0, 0%, 27%)",
                    "text-halo-blur": 0.5
                }
            },
            {
                "id": "settlement-minor-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "place-labels",
                    "mapbox:group": "Place labels, place-labels"
                },
                "source": "composite",
                "source-layer": "place_label",
                "maxzoom": 15,
                "filter": [
                    "all",
                    ["<=", ["get", "filterrank"], 3],
                    [
                        "match",
                        ["get", "class"],
                        "settlement",
                        ["match", ["get", "worldview"], ["all", "US"], true, false],
                        "disputed_settlement",
                        [
                            "all",
                            ["==", ["get", "disputed"], "true"],
                            [
                                "match",
                                ["get", "worldview"],
                                ["all", "US"],
                                true,
                                false
                            ]
                        ],
                        false
                    ],
                    [
                        "step",
                        ["zoom"],
                        true,
                        8,
                        [">=", ["get", "symbolrank"], 11],
                        10,
                        [">=", ["get", "symbolrank"], 12],
                        11,
                        [">=", ["get", "symbolrank"], 13],
                        12,
                        [">=", ["get", "symbolrank"], 15],
                        13,
                        [">=", ["get", "symbolrank"], 11],
                        14,
                        [">=", ["get", "symbolrank"], 13]
                    ]
                ],
                "layout": {
                    "icon-image": "",
                    "text-font": ["DIN Pro Regular", "Arial Unicode MS Regular"],
                    "text-offset": [
                        "step",
                        ["zoom"],
                        ["literal", [0, 0]],
                        8,
                        ["literal", [0, 0]]
                    ],
                    "text-anchor": ["step", ["zoom"], "center", 8, "center"],
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]],
                    "text-max-width": 7,
                    "text-line-height": 1.1,
                    "text-size": [
                        "interpolate",
                        ["cubic-bezier", 0.2, 0, 0.9, 1],
                        ["zoom"],
                        3,
                        [
                            "step",
                            ["get", "symbolrank"],
                            12,
                            9,
                            11,
                            10,
                            10.5,
                            12,
                            9.5,
                            14,
                            8.5,
                            16,
                            6.5,
                            17,
                            4
                        ],
                        13,
                        [
                            "step",
                            ["get", "symbolrank"],
                            23,
                            9,
                            21,
                            10,
                            19,
                            11,
                            17,
                            12,
                            16,
                            13,
                            15,
                            15,
                            13
                        ]
                    ]
                },
                "paint": {
                    "text-color": "hsl(0, 0%, 0%)",
                    "text-halo-color": "hsl(40, 53%, 100%)",
                    "text-halo-width": 1,
                    "icon-opacity": ["step", ["zoom"], 1, 8, 0],
                    "text-halo-blur": 1
                }
            },
            {
                "id": "settlement-major-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "place-labels",
                    "mapbox:group": "Place labels, place-labels"
                },
                "source": "composite",
                "source-layer": "place_label",
                "maxzoom": 15,
                "filter": [
                    "all",
                    ["<=", ["get", "filterrank"], 3],
                    [
                        "match",
                        ["get", "class"],
                        "settlement",
                        ["match", ["get", "worldview"], ["all", "US"], true, false],
                        "disputed_settlement",
                        [
                            "all",
                            ["==", ["get", "disputed"], "true"],
                            [
                                "match",
                                ["get", "worldview"],
                                ["all", "US"],
                                true,
                                false
                            ]
                        ],
                        false
                    ],
                    [
                        "step",
                        ["zoom"],
                        false,
                        8,
                        ["<", ["get", "symbolrank"], 11],
                        10,
                        ["<", ["get", "symbolrank"], 12],
                        11,
                        ["<", ["get", "symbolrank"], 13],
                        12,
                        ["<", ["get", "symbolrank"], 15],
                        13,
                        [">=", ["get", "symbolrank"], 11],
                        14,
                        [">=", ["get", "symbolrank"], 13]
                    ]
                ],
                "layout": {
                    "icon-image": "",
                    "text-font": ["DIN Pro Medium", "Arial Unicode MS Regular"],
                    "text-offset": [
                        "step",
                        ["zoom"],
                        ["literal", [0, 0]],
                        8,
                        ["literal", [0, 0]]
                    ],
                    "text-anchor": ["step", ["zoom"], "center", 8, "center"],
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]],
                    "text-max-width": 7,
                    "text-line-height": 1.1,
                    "text-size": [
                        "interpolate",
                        ["cubic-bezier", 0.2, 0, 0.9, 1],
                        ["zoom"],
                        8,
                        ["step", ["get", "symbolrank"], 18, 9, 17, 10, 15],
                        15,
                        [
                            "step",
                            ["get", "symbolrank"],
                            23,
                            9,
                            22,
                            10,
                            20,
                            11,
                            18,
                            12,
                            16,
                            13,
                            15,
                            15,
                            13
                        ]
                    ]
                },
                "paint": {
                    "text-color": "hsl(0, 0%, 0%)",
                    "text-halo-color": "hsl(40, 53%, 100%)",
                    "text-halo-width": 1,
                    "icon-opacity": ["step", ["zoom"], 1, 8, 0],
                    "text-halo-blur": 1
                }
            },
            {
                "id": "state-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "place-labels",
                    "mapbox:group": "Place labels, place-labels"
                },
                "source": "composite",
                "source-layer": "place_label",
                "minzoom": 3,
                "maxzoom": 9,
                "filter": [
                    "match",
                    ["get", "class"],
                    "state",
                    ["match", ["get", "worldview"], ["all", "US"], true, false],
                    "disputed_state",
                    [
                        "all",
                        ["==", ["get", "disputed"], "true"],
                        ["match", ["get", "worldview"], ["all", "US"], true, false]
                    ],
                    false
                ],
                "layout": {
                    "text-size": [
                        "interpolate",
                        ["cubic-bezier", 0.85, 0.7, 0.65, 1],
                        ["zoom"],
                        4,
                        ["step", ["get", "symbolrank"], 10, 6, 9.5, 7, 9],
                        9,
                        ["step", ["get", "symbolrank"], 21, 6, 16, 7, 13]
                    ],
                    "text-transform": "uppercase",
                    "text-font": ["DIN Pro Bold", "Arial Unicode MS Bold"],
                    "text-field": [
                        "step",
                        ["zoom"],
                        [
                            "step",
                            ["get", "symbolrank"],
                            ["coalesce", ["get", "name_en"], ["get", "name"]],
                            5,
                            [
                                "coalesce",
                                ["get", "abbr"],
                                ["get", "name_en"],
                                ["get", "name"]
                            ]
                        ],
                        5,
                        ["coalesce", ["get", "name_en"], ["get", "name"]]
                    ],
                    "text-letter-spacing": 0.15,
                    "text-max-width": 6
                },
                "paint": {
                    "text-color": "hsl(0, 0%, 0%)",
                    "text-halo-color": "hsl(40, 53%, 100%)",
                    "text-halo-width": 1
                }
            },
            {
                "id": "country-label",
                "type": "symbol",
                "metadata": {
                    "mapbox:featureComponent": "place-labels",
                    "mapbox:group": "Place labels, place-labels"
                },
                "source": "composite",
                "source-layer": "place_label",
                "minzoom": 1,
                "maxzoom": 10,
                "filter": [
                    "match",
                    ["get", "class"],
                    "country",
                    ["match", ["get", "worldview"], ["all", "US"], true, false],
                    "disputed_country",
                    [
                        "all",
                        ["==", ["get", "disputed"], "true"],
                        ["match", ["get", "worldview"], ["all", "US"], true, false]
                    ],
                    false
                ],
                "layout": {
                    "icon-image": "",
                    "text-field": ["coalesce", ["get", "name_en"], ["get", "name"]],
                    "text-line-height": 1.1,
                    "text-max-width": 6,
                    "text-font": ["DIN Pro Medium", "Arial Unicode MS Regular"],
                    "text-offset": ["literal", [0, 0]],
                    "text-size": [
                        "interpolate",
                        ["cubic-bezier", 0.2, 0, 0.7, 1],
                        ["zoom"],
                        1,
                        ["step", ["get", "symbolrank"], 11, 4, 9, 5, 8],
                        9,
                        ["step", ["get", "symbolrank"], 22, 4, 19, 5, 17]
                    ]
                },
                "paint": {
                    "icon-opacity": [
                        "step",
                        ["zoom"],
                        ["case", ["has", "text_anchor"], 1, 0],
                        7,
                        0
                    ],
                    "text-color": "hsl(0, 0%, 0%)",
                    "text-halo-color": [
                        "interpolate",s
                        ["linear"],
                        ["zoom"],
                        2,
                        "hsla(40, 53%, 100%, 0.75)",
                        3,
                        "hsl(40, 53%, 100%)"
                    ],
                    "text-halo-width": 1.25
                }
            },
            {
                "id": "'''+SourceLayerId+'''",
                "type": "circle",
                "source": "composite",
                "source-layer": "'''+SourceLayerName+'''",
                "layout": {},
                "paint": {
                    "circle-radius": [
                        "interpolate",
                        ["linear"],
                        ["get", "'''+Variable+'''"],
                        1,
                        8,
                        5,
                        12
                    ],
                    "circle-color": [
                        "interpolate",
                        ["linear"],
                        ["get", "'''+Variable+'''"],
                        1,
                        "hsl('''+OneColorH+''', '''+OneColorS+'''%, '''+OneColorL+'''%)",
                        5,
                        "hsl('''+TwoColorH+''', '''+TwoColorS+'''%, '''+TwoColorL+'''%)"
                    ],
                    "circle-opacity": [
                        "interpolate",
                        ["linear"],
                        ["get", "'''+Variable+'''"],
                        1,
                        0.7,
                        5,
                        1
                    ]
                }
            }
        ],
        "created": "2020-12-03T18:04:12.799Z",
        "modified": "2020-12-03T18:10:46.459Z",
        "id": "cki95ew494au719ufyzwo9v7j",
        "owner": "carmela-cucuzzella",
        "visibility": "private",
        "draft": false
    }'''

    return Text

def GetData(ExcelPath,ExitFolder):
    wb = load_workbook(filename = ExcelPath)
    print(wb.sheetnames)
    SheetName='Data for Styles'
    print(wb[SheetName]['A1'].value)
    print("............")
    print("............")


    CellCenDeg_H1="M2"
    CellCenClo_H1="M3"
    CellCenEig_H1="M4"
    CellWeight_H1="M5"
    CellCenDeg_S1="N2"
    CellCenClo_S1="N3"
    CellCenEig_S1="N4"
    CellWeight_S1="N5"
    CellCenDeg_L1="O2"
    CellCenClo_L1="O3"
    CellCenEig_L1="O4"
    CellWeight_L1="O5"
    CellCenDeg_H2="P2"
    CellCenClo_H2="P3"
    CellCenEig_H2="P4"
    CellWeight_H2="P5"
    CellCenDeg_S2="Q2"
    CellCenClo_S2="Q3"
    CellCenEig_S2="Q4"
    CellWeight_S2="Q5"
    CellCenDeg_L2="R2"
    CellCenClo_L2="R3"
    CellCenEig_L2="R4"
    CellWeight_L2="R5"

    ValueCenDeg_H1=wb[SheetName][CellCenDeg_H1].value
    ValueCenClo_H1=wb[SheetName][CellCenClo_H1].value
    ValueCenEig_H1=wb[SheetName][CellCenEig_H1].value
    ValueWeiht__H1=wb[SheetName][CellWeight_H1].value

    ValueCenDeg_S1=wb[SheetName][CellCenDeg_S1].value
    ValueCenClo_S1=wb[SheetName][CellCenClo_S1].value
    ValueCenEig_S1=wb[SheetName][CellCenEig_S1].value
    ValueWeiht__S1=wb[SheetName][CellWeight_S1].value

    ValueCenDeg_L1=wb[SheetName][CellCenDeg_L1].value
    ValueCenClo_L1=wb[SheetName][CellCenClo_L1].value
    ValueCenEig_L1=wb[SheetName][CellCenEig_L1].value
    ValueWeiht__L1=wb[SheetName][CellWeight_L1].value

    ValueCenDeg_H2=wb[SheetName][CellCenDeg_H2].value
    ValueCenClo_H2=wb[SheetName][CellCenClo_H2].value
    ValueCenEig_H2=wb[SheetName][CellCenEig_H2].value
    ValueWeiht__H2=wb[SheetName][CellWeight_H2].value

    ValueCenDeg_S2=wb[SheetName][CellCenDeg_S2].value
    ValueCenClo_S2=wb[SheetName][CellCenClo_S2].value
    ValueCenEig_S2=wb[SheetName][CellCenEig_S2].value
    ValueWeiht__S2=wb[SheetName][CellWeight_S2].value

    ValueCenDeg_L2=wb[SheetName][CellCenDeg_L2].value
    ValueCenClo_L2=wb[SheetName][CellCenClo_L2].value
    ValueCenEig_L2=wb[SheetName][CellCenEig_L2].value
    ValueWeiht__L2=wb[SheetName][CellWeight_L2].value


    print("ValueCenDeg_H1",ValueCenDeg_H1)
    print("ValueCenClo_H1",ValueCenClo_H1)
    print("ValueCenEig_H1",ValueCenEig_H1)
    print("ValueCenDeg_S1",ValueCenDeg_S1)
    print("ValueCenClo_S1",ValueCenClo_S1)
    print("ValueCenEig_S1",ValueCenEig_S1)
    print("ValueCenDeg_L1",ValueCenDeg_L1)
    print("ValueCenClo_L1",ValueCenClo_L1)
    print("ValueCenEig_L1",ValueCenEig_L1)
    print("ValueCenDeg_H2",ValueCenDeg_H2)
    print("ValueCenClo_H2",ValueCenClo_H2)
    print("ValueCenEig_H2",ValueCenEig_H2)
    print("ValueCenDeg_S2",ValueCenDeg_S2)
    print("ValueCenClo_S2",ValueCenClo_S2)
    print("ValueCenEig_S2",ValueCenEig_S2)
    print("ValueCenDeg_L2",ValueCenDeg_L2)
    print("ValueCenClo_L2",ValueCenClo_L2)
    print("ValueCenEig_L2",ValueCenEig_L2)





    CellHasData=True
    RowCount=1
    while(CellHasData==True):
        RowCount=RowCount+1
        print("............................................................"+str(RowCount)+"............................................................")
        CellValue='A'+str(RowCount)
        CellValue=wb[SheetName][CellValue].value
        # b=input("Stop.")
        if str(CellValue)=='None':
            # print("Kill me!!!")
            CellHasData=False
        else:
            print("CellValue",CellValue,type(CellValue))
            CellCity='A'+str(RowCount)
            CellTitle='B'+str(RowCount)
            CellTileSet='C'+str(RowCount)
            CellCoordX='D'+str(RowCount)
            CellCoordY='E'+str(RowCount)
            CellURLtileset='F'+str(RowCount)
            CellSourceLayerId='G'+str(RowCount)
            CellSourceLayerName='H'+str(RowCount)
            CellVariable='I'+str(RowCount)

            ValueCity=wb[SheetName][CellCity].value
            ValueTitle=wb[SheetName][CellTitle].value
            ValueTileSet=wb[SheetName][CellTileSet].value
            ValueCoordX=wb[SheetName][CellCoordX].value
            ValueCoordY=wb[SheetName][CellCoordY].value
            ValueURLtileset=wb[SheetName][CellURLtileset].value
            ValueSourceLayerId=wb[SheetName][CellSourceLayerId].value
            ValueSourceLayerName=wb[SheetName][CellSourceLayerName].value
            ValueVariable=wb[SheetName][CellVariable].value

            # print("ValueCity",ValueCity,end="\t")
            # print("ValueTitle",ValueTitle,end="\t")
            # print("ValueTileSet",ValueTileSet,end="\t")
            # print("ValueCoordX",ValueCoordX,end="\t")
            # print("ValueCoordY",ValueCoordY,end="\t")
            # print("ValueURLtileset",ValueURLtileset,end="\t")
            # print("ValueSourceLayerId",ValueSourceLayerId,end="\t")
            # print("ValueSourceLayerName",ValueSourceLayerName,end="\t")
            # print("ValueVariable",ValueVariable,end="\t")
            # print("\n"*2)
            
            MapName=ValueTitle
            MapCode='cki95ew494ajijijijijijijiji'
            TileSet=ValueTileSet
            CoordMapX=ValueCoordX
            CoordMapY=ValueCoordY
            URLtileset=ValueURLtileset
            SourceLayerId=ValueSourceLayerId
            SourceLayerName=ValueSourceLayerName
            Variable=ValueVariable

            ExitPath=ExitFolder+"\Map_"+ValueTitle+".json"
            fw=open(ExitPath,"w")

            if ValueVariable =="CatCenDeg":
                OneColorH=ValueCenDeg_H1
                OneColorS=ValueCenDeg_S1
                OneColorL=ValueCenDeg_L1
                TwoColorH=ValueCenDeg_H2
                TwoColorS=ValueCenDeg_S2
                TwoColorL=ValueCenDeg_L2
            if ValueVariable =="CatClossnes":
                OneColorH=ValueCenClo_H1
                OneColorS=ValueCenClo_S1
                OneColorL=ValueCenClo_L1
                TwoColorH=ValueCenClo_H2
                TwoColorS=ValueCenClo_S2
                TwoColorL=ValueCenClo_L2
            if ValueVariable =="CatEigen":
                OneColorH=ValueCenEig_H1
                OneColorS=ValueCenEig_S1
                OneColorL=ValueCenEig_L1
                TwoColorH=ValueCenEig_H2
                TwoColorS=ValueCenDeg_S2
                TwoColorL=ValueCenEig_L2
            if ValueVariable =="weight":
                OneColorH=ValueWeiht__H1
                OneColorS=ValueWeiht__S1
                OneColorL=ValueWeiht__L1
                TwoColorH=ValueWeiht__H2
                TwoColorS=ValueWeiht__S2
                TwoColorL=ValueWeiht__L2
            if ValueVariable =="NA":
                OneColorH='144'
                OneColorS='96'
                OneColorL='62'
                TwoColorH='144'
                TwoColorS='92'
                TwoColorL='25'

            print("MapName",MapName)
            print("MapCode",MapCode)
            print("TileSet",TileSet)
            print("CoordMapX",CoordMapX)
            print("CoordMapY",CoordMapY)
            print("URLtileset",URLtileset)
            print("SourceLayerId",SourceLayerId)
            print("SourceLayerName",SourceLayerName)
            print("Variable",Variable)
            print("OneColorH",OneColorH)
            print("OneColorS",OneColorS)
            print("OneColorL",OneColorL)
            print("TwoColorH",TwoColorH)
            print("TwoColorS",TwoColorS)
            print("TwoColorL",TwoColorL)

            Layout=TextSource()
            fw.write(Layout)
            fw.close()



ExcelPath="E:\OneDrive - Concordia University - Canada\RA-CAMM\Software\CAMMM-Soft-Tool_V1.1\Control.xlsx"
ExitFolder="E:\OneDrive - Concordia University - Canada\RA-CAMM\Software\CAMMM-Soft-Tool_V1.1\MapBoxOutput"


GetData(ExcelPath=ExcelPath,ExitFolder=ExitFolder)
# print(TextSource())


# print(Text)

# Path="E:\OneDrive - Concordia University - Canada\RA-CAMM\Software\CAMMM-Soft-Tool_V1.1\Scripts\Closnes-Montreal-Node(cki3lqovc29i31amdcurzd6aw)\Closnes-Quebec-Direct(cki95ew494au719ufyzwo9v7j)\Style_Exit3.json"
# fw=open(Path,'w')
# fw.write(Text)
# fw.close()zs