import openpyxl
from openpyxl import load_workbook



def ReadHowManyRecordsWeHave(WorkSheet):
    EmptyCell=False
    # print("--------------",WorkSheet.cell(row=1, column=1).value)
    DataPairs={}
    RowVal=0
    while(EmptyCell==False):
        RowVal=RowVal+1
        if WorkSheet.cell(row=RowVal, column=1).value == None:
            print("Fin")
            EmptyCell=True
        else:
            # print("***",WorkSheet.cell(row=RowVal, column=1).value)
            Index=WorkSheet.cell(row=RowVal, column=1).value
            Name=WorkSheet.cell(row=RowVal, column=2).value
            DataPairs[Name]=Index
            print("***",Name,Index)
    return RowVal-1,DataPairs


def WriteToMainExcel(Data):
    PathExcel="E:\OneDrive - Concordia University - Canada\RA-CAMM\Software\CAMMM-Soft-Tool_V1.1\Database.xlsx"
    wb = load_workbook(filename = PathExcel)
    print (wb.sheetnames, type(wb.sheetnames))
    sheets=list(wb.sheetnames)
    ListOfCities=wb[sheets[0]]
    print("ListOfCities",ListOfCities,type(ListOfCities))
    NumberOfRecordsInCode,DictwithKeyofCities=ReadHowManyRecordsWeHave(WorkSheet=ListOfCities)
    print("NumberOfRecordsInCode",NumberOfRecordsInCode)
    Check=False
    if Data["Agency"] in wb.sheetnames:
        print("Record is in sheet list")
        Check=True
    if Data["Agency"] in ListOfCities:
        print("Record is in sheet list")
        Check=True


    if Check==False: 
        print("Number of stuff,",len(DictwithKeyofCities.keys()))
        IndexCell="A"+str(len(DictwithKeyofCities.keys())+1)
        AgencyCell="B"+str(len(DictwithKeyofCities.keys())+1)
        print("IndexCell",IndexCell)
        print("AgencyCell",AgencyCell)
        ListOfCities[IndexCell]=len(DictwithKeyofCities.keys())+1
        ListOfCities[AgencyCell]=Data["Agency"]

        print("New record",Data["Agency"])
        ws1 = wb.create_sheet(Data["Agency"])
        ws1["A1"]="Number of transport systems"
        ws1["A2"]="Number of train stations"
        ws1["A3"]="Number of metro stations"
        ws1["A4"]="Number of bus stops"
        ws1["A5"]="Number of boroughs"
        ws1["A6"]="Area"
        ws1["A7"]="Number of train stations"
        ws1["A8"]="Density"

        ws1["B1"]=Data["Number_of_transport_systems"]
        ws1["B2"]=Data["Number_of_train_stations"]
        ws1["B3"]=Data["Number_of_metro_stations"]
        ws1["B4"]=Data["Number_of_bus_stops"]
        # ws1["B5"]=
        # ws1["B6"]=
        # ws1["B7"]=
        ws1["B8"]="=B7/B6"
        wb.save(PathExcel)






if __name__ == "__main__":
    Data={"Agency":"Agency A10",
    "Number_of_transport_systems":20,
    "Number_of_train_stations":340,
    "Number_of_metro_stations":2354,
    "Number_of_bus_stops":22486}
    WriteToMainExcel(Data=Data)