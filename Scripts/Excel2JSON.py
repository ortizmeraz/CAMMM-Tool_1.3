import openpyxl
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfile
from tkinter import filedialog
from tkinter.filedialog import askdirectory
from tkinter import messagebox
from shutil import copyfile
import tkinter
from functools import partial


def Example():
    Dict={"City": {
        "Montreal": {"name": "Montréal","NumTransportSystem": 3,"NumBusStops": 65666,"NumBusLines":255,"NumRailStations": 123,"NumMetroStations": 50,"NumBoroughs": 19,"AreaSqKm": 431.5,"PopulationMillion": 1.78,"DensityPersonSqKm": 4115,"DirectStyleURL": "ckmdts11g14q417pe6b2r7ytx","NodeStyleURL": "ckmduljdi15jd17pdh44qtoh8","DirectLayers": [    "montreal_D_CD",    "montreal_D_CL"],"NodeLayers": [    "montreal_N_CD",    "montreal_N_CL"],"Coords": [    -73.71,    45.53],"Zoom": 10.45},
        "Vienna": {"name": "Vienna","NumTransportSystem": 6,"NumBusStops": 78888,"NumBusLines":256,"NumRailStations": 456,"NumMetroStations": 35,"NumBoroughs": 20,"AreaSqKm": 414.6,"PopulationMillion": 1.8,"DensityPersonSqKm": 4341,"DirectStyleURL": "ckmdv9k6p0y6617qwpyikvpms","NodeStyleURL": "ckmdvdgrv16cx17pljoyexm36","DirectLayers": [    "vienna_D_CD",    "vienna_D_CL"],"NodeLayers": [    "vienna_N_CD",    "vienna_N_CL"],"Coords": [    16.39,    48.20],"Zoom": 11},
        "Barcelona": {"name": "Barcelona","NumTransportSystem": 6,"NumBusStops": 44444,"NumBusLines":257,"NumRailStations": 36,"NumMetroStations": 135,"NumBoroughs": 1,"AreaSqKm": 101.9,"PopulationMillion": 5.575,"DensityPersonSqKm": 54710,"DirectStyleURL": "ckmdvikhj16k118rv36unif4h","NodeStyleURL": "ckmdvr9px0yne17qw1abv6rqx","DirectLayers": [    "barcelona_D_CD",    "barcelona_D_CL",    "barcelona_D_EV"],"NodeLayers": [    "barcelona_N_CD",    "barcelona_N_CL",    "barcelona_N_EV"],"Coords": [    2.13,    41.38],"Zoom": 11.81},
        "Budapest": {"name": "Budapest","NumTransportSystem": 2,"NumBusStops": 3333333,"NumBusLines":258,"NumRailStations": 66,"NumMetroStations": 222,"NumBoroughs": 15,"AreaSqKm": 525.2,"PopulationMillion": 1.75,"DensityPersonSqKm": 3388,"DirectStyleURL": "ckmdvy9vn0yud17qwen6l9jj8","NodeStyleURL": "ckmdw5kz7171117pddh6zl0bj","DirectLayers": [    "budapest_D_CD",    "budapest_D_CL"],"NodeLayers": [    "budapest_N_CD",    "budapest_N_CL"],"Coords": [    19.12,    47.48],"Zoom": 10},
        "Quebec": {"name": "Quebec","NumTransportSystem": 4,"NumBusStops": 777777,"NumBusLines":259,"NumRailStations": 77,"NumMetroStations": 222,"NumBoroughs": 10,"AreaSqKm": 234.5,"PopulationMillion": 0.57,"DensityPersonSqKm": 1648,"DirectStyleURL": "ckmdwizn553xs17naq3j69fy6","NodeStyleURL": "ckmdwsyti17nu17pdw6orgej8","DirectLayers": [    "quebec_D_CD",    "quebec_D_CL"],"NodeLayers": [    "quebec_N_CD",    "quebec_N_CL"],"Coords": [    -71.26,    46.78],"Zoom": 11.3}
    }}
    return Dict

def ReadHowManyRecordsWeHave(WorkSheet):
    EmptyCell=False
    RowVal=2
    ListOhCities=[]
    while(EmptyCell==False):
        RowVal=RowVal+1
        if WorkSheet.cell(row=RowVal, column=1).value == None:
            print("Fin")
            EmptyCell=True
        else:
            Index=WorkSheet.cell(row=RowVal, column=1).value
            Name=WorkSheet.cell(row=RowVal, column=2).value
            ListOhCities.append(Index)
            print("***",Name,Index)
    return len(ListOhCities)

def CreateListRows(NumRecords):
    ListRows=[]
    for i in range(3,(NumRecords+3)):
        print("-----------",i)
        ListRows.append(i)
    return ListRows

def GetColumns():
    Columns={}
    Columns["City"]="A"
    Columns["name"]="B"
    Columns["NumTransportSystem"]="C"
    Columns["NumBusStops"]="D"
    Columns["NumBusLines"]="E"
    Columns["NumRailStations"]="F"
    Columns["NumMetroStations"]="G"
    Columns["NumBoroughs"]="H"
    Columns["AreaSqKm"]="I"
    Columns["PopulationMillion"]="J"
    Columns["DensityPersonSqKm"]="K"
    Columns["DirectStyleURL"]="L"
    Columns["NodeStyleURL"]="M"
    Columns["DirectLayers"]="N"
    Columns["NodeLayers"]="O"
    Columns["Lat"]="P"
    Columns["Lon"]="Q"
    Columns["Zoom"]="R"
    return Columns


def WriteFile(DataSet,ExitPath):
    print(DataSet,ExitPath)
    fw=open(file=ExitPath,mode="w")
    print("Yeah",fw)


    

def ReadExcelTable(NumRecords,WorkSheet):
    DataSet={"City":{}}
    # RowVal=RowVal+1
    ListofRows=CreateListRows(NumRecords)
    DictColumns=GetColumns()
    for row in ListofRows:
        DataSet["City"][WorkSheet["A"+str(row)].value]={}
        # print(row,"\t",WorkSheet["A"+str(row)].value,"\t",end="\t")
        for column in DictColumns.keys():
            Cell=DictColumns[column]+str(row)
            # print(Cell,end="\t")
            # print(Cell,WorkSheet[Cell].value,end="\t")
            # print("CityKey",WorkSheet["A"+str(row)].value,end="\t")
            # print("DictColumns",column,end="\t")
            if DictColumns[column] in ("N","O"):
                Text=WorkSheet[Cell].value
                print("original text",Text)
                ListLayers=Text.split(",")
                print("City:",WorkSheet["A"+str(row)].value)
                print("Attribute:",column)
                print("ListLayers",ListLayers)
                DataSet["City"][WorkSheet["A"+str(row)].value][column]=ListLayers
                # b=input()
                # print(b)
            elif DictColumns[column] == "P":
                Coords=[WorkSheet[Cell].value]
            elif DictColumns[column]=="Q":
                Coords.append(WorkSheet[Cell].value)
                DataSet["City"][WorkSheet["A"+str(row)].value]["Coords"]=Coords
            else:
                DataSet["City"][WorkSheet["A"+str(row)].value][column]=WorkSheet[Cell].value
        print()
    # PrintCheck(DataSet=DataSet)
    WriteFile(DataSet=DataSet,ExitPath="sssssssssssss")

def PrintCheck(DataSet):
    print(DataSet)
    print("set one....................................................")
    for key in DataSet["City"].keys():
        print ("key:",key)
        for CityKey in DataSet["City"][key].keys():
            print(CityKey,end=":")
        print()
    print()
    print("set two....................................................")
    for key in Example()["City"].keys(): 
        print ("key:",key)
        for CityKey in Example()["City"][key].keys():
            # print(CityKey,end=":")
            print("Sample",CityKey,Example()["City"][key][CityKey],type(Example()["City"][key][CityKey]))
            print("Data",CityKey,DataSet["City"][key][CityKey],type(DataSet["City"][key][CityKey]))
            print()

def OpenFile(PathExcel):
    
    wb = load_workbook(filename = PathExcel)
    print (wb.sheetnames, type(wb.sheetnames))
    sheets=list(wb.sheetnames)
    print(sheets)
    DataSheet=wb[sheets[0]]
    print(DataSheet)
    NumberOfRecordsInCode=ReadHowManyRecordsWeHave(WorkSheet=DataSheet)
    print("NumberOfRecordsInCode",NumberOfRecordsInCode)
    ReadExcelTable(NumRecords=NumberOfRecordsInCode,WorkSheet=DataSheet)

def Run(SourceData,JSONData):
    PathExcel=SourceData.GivePath()
    PathExit=JSONData.GivePath()
    print("VAriable PathExcel:",PathExcel)
    print("VAriable PathExit:",PathExit)

class PathFrame():
    def __init__(self,Frame,Path,CoordX,CoordY,Title,FileType):
        self.Frame=Frame
        self.Path=Path
        self.CoordX=CoordX
        self.CoordY=CoordY
        self.Title=Title
        self.FileType=FileType
        
    def getPathRead(self):
        Path = askopenfilename(initialdir=r"E:\GitHub\CAMMM-Web-Tool-1",
                    filetypes =(self.FileType,("All Files","*.*")),
                    title = "Choose a file."
                    )
        self.Path=Path        
        print(self.Path,type(self.Path))
        self.EntryElement.insert(0,self.Path)
        
    def getPathWrite(self):
        Path = asksaveasfile(initialdir=r"E:\GitHub\CAMMM-Web-Tool-1",
                    filetypes =(self.FileType,("All Files","*.*")),
                    title = "Choose a file."
                    )
        self.Path=Path.name  
        # print(dir(self.Path))   
        # for i in list(dir(self.Path)):
        #      print(self.Path.i)
        # print(self.Path,type(self.Path))
        self.EntryElement.insert(0,self.Path)

    def MakeGuiElementsRead(self):
        self.FrameElement=ttk.LabelFrame(self.Frame,height = 55, width = 700,text =self.Title)
        self.FrameElement.place(x=self.CoordX, y=self.CoordY, anchor=W)

        self.EntryElement= ttk.Entry(self.FrameElement,width=100)
        self.EntryElement.place(x=5, y=9, anchor=W)
        self.EntryElement.delete(0, END)
        self.EntryElement.insert(0,"")

        self.button=ttk.Button(self.FrameElement,text="Open")
        self.button.config(command=self.getPathRead)
        self.button.place(x=615, y=10, anchor=W)


    def MakeGuiElementsWrite(self):
        self.FrameElement=ttk.LabelFrame(self.Frame,height = 55, width = 700,text =self.Title)
        self.FrameElement.place(x=self.CoordX, y=self.CoordY, anchor=W)

        self.EntryElement= ttk.Entry(self.FrameElement,width=100)
        self.EntryElement.place(x=5, y=9, anchor=W)
        self.EntryElement.delete(0, END)
        self.EntryElement.insert(0,"")

        self.button=ttk.Button(self.FrameElement,text="Open")
        self.button.config(command=self.getPathWrite)
        self.button.place(x=615, y=10, anchor=W)



    def GivePath(self):
        return self.Path



def GUI(root,PathExcel):
    root.title("Concordia Research Chair in Integrated Design, Ecology And Sustainability for the Built Environment")
    FrameMaster =ttk.Frame(root)   
    FrameMaster =ttk.Frame(root)
    FrameMaster.config(height=300,width=750) # Configura altura y ancho
    FrameMaster.config(relief= RIDGE)        # Tipod de frame
    FrameMaster.config(padding=(30, 15))     # Acolchonado de frame
    FrameMaster.pack()                       # ANade frame

    SourceData=PathFrame(Frame=FrameMaster,Path="",CoordX=1,CoordY=55,Title= "Excel File",FileType=("Excel 365 Files", "*.xlsx"))
    JSONData=PathFrame(Frame=FrameMaster,Path="",CoordX=1,CoordY=155,Title= "JSON File",FileType=("JSON", "*.json"))

    SourceData.MakeGuiElementsRead()
    JSONData.MakeGuiElementsWrite()
    
    buttonRun=ttk.Button(FrameMaster,text="Open")
    buttonRun.config(command=partial(Run,SourceData,JSONData))
    buttonRun.place(x=10, y=255, anchor=W)

    root.mainloop()

if __name__ == "__main__":

    # OpenFile(PathExcel=r"E:\GitHub\CAMMM-Web-Tool-1\DatabaseCitys.xlsx")
    root=Tk()
    PathExcel =""
    GUI(root,PathExcel)
